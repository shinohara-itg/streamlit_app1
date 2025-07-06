# streamlit_combined_app.py

import streamlit as st
import os
from tempfile import NamedTemporaryFile
from exe01_excel_in_csv_out import process_excel  # ステップ2の関数（excel_processor.py に保存）
from exe02_Q_analysis import run_q_analysis      # ステップ4の関数（q_analysis.py に保存）
from exe03_txt_review import run_txt_review      # ステップ6の関数（q_analysis.py に保存）
from exe04_summry_from_outfile import run_summary_from_outfile
from exe06_commentback_to_chartfile import write_txt_to_excel


# タブの定義
tab1, tab2, tab3, tab4, tab5, tab6, tab7= st.tabs([
    "STEP1", # データ生成
    "STEP2", # LLM分析
    "STEP3", # 校正
    "STEP4", # データレビュー&LR用データ抽出
    "STEP5", # サマリー生成
    "STEP6", # サマリー内容レビュー
    "STEP7"]) # サマリーのエクセル出力

# ----------------------------------------
#    "STEP1", # データ生成
# ----------------------------------------
with tab1:
    st.markdown(
    """
    ### STEP1:データ生成  
    <span style='font-size:10pt;'>LE出力データからAIに渡すデータを作ります。</span>
    """,
    unsafe_allow_html=True
)

    uploaded_file = st.file_uploader("Excelファイルをアップロード", type=["xlsx"], key="excel_uploader")
    base_dir = st.text_input("出力先のベースフォルダパス（例：C:/Users/.../05 Moving Estimate service）")

    if st.button("STEP1の処理を実行"):
        if uploaded_file is None:
            st.error("Excelファイルを選択してください。")
        elif not base_dir:
            st.error("ベースフォルダを入力してください。")
        else:
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_file.write(uploaded_file.getbuffer())
                temp_excel_path = tmp_file.name

            try:
                process_excel(temp_excel_path, base_dir)
                st.success("STEP1：処理完了！STEP2へ進んでください！")

                # セッション状態にパス保存
                st.session_state["base_dir"] = base_dir
                st.session_state["uploaded_file_path"] = temp_excel_path
                print(f"📄 Excel{temp_excel_path}")

            except Exception as e:
                st.error(f"エラーが発生しました：{e}")

# ----------------------------------------
#     "STEP2", # LLM分析
# ----------------------------------------
with tab2:
    st.header("STEP2: LLM分析　クロス集計とプロンプトをAIに渡して分析します。")

    if "base_dir" in st.session_state:
        base_dir = st.session_state["base_dir"]
        st.info(f"STEP2で指定されたフォルダ： {base_dir}")

        input_csv_dir = os.path.join(base_dir, "1 datatable")
        output_txt_dir = os.path.join(base_dir, "2 output")
        temp_dir = os.path.join(base_dir, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        temp_prompt_path = os.path.join(temp_dir, "temp_prompt.txt")

    else:
        st.warning("STEP1を先に実行してください。")
        input_csv_dir = ""
        output_txt_dir = ""
        temp_prompt_path = "temp_prompt.txt"  # フォールバック

    prompt_file = st.file_uploader("共通プロンプトファイル（.txt）をアップロード", type=["txt"], key="prompt_uploader")

    prompt_text = ""
    if prompt_file:
        prompt_text = prompt_file.read().decode("utf-8")
        edited_prompt_text = st.text_area("プロンプト内容（編集可）", value=prompt_text, height=900, key="editable_prompt")

        if st.button("修正したプロンプトを保存"):
            try:
                with open(temp_prompt_path, "w", encoding="utf-8") as f:
                    f.write(edited_prompt_text)
                st.success(f"プロンプトを保存しました：{temp_prompt_path}")
            except Exception as e:
                st.error(f"保存中にエラーが発生しました：{e}")

        if st.button("LLM分析を実行"):
            if not os.path.exists(temp_prompt_path):
                st.error("まずプロンプトを保存してください。")
            else:
                try:
                    with st.spinner("LLMで分析中..."):
                        run_q_analysis(temp_prompt_path, input_csv_dir, output_txt_dir)
                        st.success("STEP2：分析完了！ STEP3に進んでください！")
                except Exception as e:
                    st.error(f"エラーが発生しました：{e}")


# ----------------------------------------
#     "STEP3", # 校正
# ----------------------------------------
# 追加する新しいタブ
with tab3:
    st.header("STEP3:校正　表記統一を行います。")

    if "base_dir" in st.session_state:
        base_dir = st.session_state["base_dir"]
        input_txt_dir = os.path.join(base_dir, "2 output")
        output_txt_dir = os.path.join(base_dir, "3 output ver2")
        temp_dir = os.path.join(base_dir, "temp")
        os.makedirs(temp_dir, exist_ok=True)
    else:
        st.warning("STEP2を先に実行してください。")
        base_dir = ""
        input_txt_dir = ""
        output_txt_dir = ""

    # --- プロンプトファイルのアップロード＆編集表示 ---
    prompt_file_review = st.file_uploader("表記統一用プロンプトファイル（.txt）", type=["txt"], key="prompt_uploader_review")

    if prompt_file_review:
        prompt_text_review = prompt_file_review.read().decode("utf-8")
        edited_prompt_review = st.text_area("プロンプト内容（編集可）", value=prompt_text_review, height=900, key="editable_prompt_review")

        if st.button("STEP3 表記統一を実行"):
            try:
                with st.spinner("表記統一を実行中..."):
                    run_txt_review(base_dir, edited_prompt_review)
                    st.success("STEP3：表記統一が完了しました！STEP4へ進んでください！")
            except Exception as e:
                st.error(f"エラーが発生しました：{e}")


# ----------------------------------------
#     "STEP4", # データレビュー&LR用データ抽出
# ----------------------------------------
with tab4:
    st.header("STEP4: クロス集計分析結果のレビュー。OKならLR用データを抽出！")

    if "base_dir" in st.session_state:
        base_dir = st.session_state["base_dir"]
        output_dir = os.path.join(base_dir, "2 output")
        file_list = [f for f in os.listdir(output_dir) if f.endswith(".txt")]

        if not file_list:
            st.info("レビュー対象のテキストファイルが見つかりません。")
        else:
            selected_file = st.selectbox("確認・修正するファイルを選択", file_list)
            selected_path = os.path.join(output_dir, selected_file)

            with open(selected_path, "r", encoding="utf-8") as f:
                file_content = f.read()

            updated_text = st.text_area("内容を確認・必要に応じて修正してください", value=file_content, height=800)

            if st.button("修正内容を保存", key="save_review_text"):
                try:
                    with open(selected_path, "w", encoding="utf-8") as f:
                        f.write(updated_text)
                    st.success("修正内容を保存しました！")
                except Exception as e:
                    st.error(f"保存中にエラーが発生しました：{e}")

        # --- 追加機能：全ファイルから【データの特徴】抽出処理 ---
        if st.button("LR用データを保存", key="extract_feature_btn"):
            try:
                input_folder = os.path.join(base_dir, "3 output ver2")
                output_folder = os.path.join(base_dir, "4 output for LR")
                os.makedirs(output_folder, exist_ok=True)

                count = 0
                for file_name in os.listdir(input_folder):
                    if not file_name.endswith(".txt"):
                        continue

                    input_path = os.path.join(input_folder, file_name)
                    output_path = os.path.join(output_folder, file_name)

                    with open(input_path, "r", encoding="utf-8") as f:
                        lines = f.readlines()

                    in_target_section = False
                    extracted_lines = []

                    for line in lines:
                        stripped = line.strip()
                        if stripped.startswith("【データの特徴】"):
                            in_target_section = True
                            continue
                        if in_target_section:
                            if stripped.startswith("【"):  # 次セクション
                                break
                            if stripped.startswith("・"):
                                extracted_lines.append(stripped)

                    with open(output_path, "w", encoding="utf-8") as f:
                        f.write("\n".join(extracted_lines))

                    count += 1

                st.success(f"✅ {count} 件のファイルから【データの特徴】を抽出・保存しました。")
            except Exception as e:
                st.error(f"抽出処理でエラーが発生しました：{e}")

    else:
        st.warning("STEP2を先に実行してください。")


# ----------------------------------------
#     "STEP5", # サマリー生成
# ----------------------------------------

with tab5:
    st.header("STEP5:サマリー生成　クロス集計分析の結果からサマリーを作成します。")

    if "base_dir" in st.session_state:
        base_dir = st.session_state["base_dir"]
        st.info(f"STEP3までの出力フォルダ： {base_dir}/3 output ver2 を使用")

        # --- プロンプトアップロード・編集 ---
        prompt_file = st.file_uploader("サマリープロンプトファイル（.txt）をアップロード", type=["txt"], key="summary_prompt_uploader")

        if prompt_file:
            prompt_text = prompt_file.read().decode("utf-8")
            edited_prompt_text = st.text_area("プロンプト内容（編集可）", value=prompt_text, height=800, key="summary_prompt_text")

            if st.button("サマリーを実行"):
                try:
                    with st.spinner("LLMで要約中..."):
                        run_summary_from_outfile(base_dir, edited_prompt_text)
                        st.success("step5:要約が完了しました！STEP6へ進んでください！")
                except Exception as e:
                    st.error(f"エラーが発生しました：{e}")
    else:
        st.warning("STEP2を先に実行してください。")



# ----------------------------------------
#     "STEP6", # サマリー内容レビュー
# ----------------------------------------
with tab6:
    st.header("STEP6:サマリー内容のレビュー・編集")

    if "base_dir" in st.session_state:
        base_dir = st.session_state["base_dir"]
        summary_path = os.path.join(base_dir, "4 summury", "all_summary.txt")

        if os.path.exists(summary_path):
            with open(summary_path, "r", encoding="utf-8") as f:
                summary_text = f.read()

            updated_text = st.text_area("サマリー内容（編集可）", value=summary_text, height=900, key="editable_summary_text")

            if st.button("修正内容を保存", key="save_summary_btn"):
                try:
                    with open(summary_path, "w", encoding="utf-8") as f:
                        f.write(updated_text)
                    st.success("修正内容を保存しました！")
                except Exception as e:
                    st.error(f"保存中にエラーが発生しました：{e}")
        else:
            st.warning("サマリーファイル（all_summary.txt）がまだ生成されていません。STEP7を先に実行してください。")
    else:
        st.warning("STEP2を先に実行してください。")

# ----------------------------------------
#     "STEP7" # サマリーのエクセル出力
# ----------------------------------------
from openpyxl import load_workbook
import os
import tempfile
import streamlit as st

def write_txt_to_excel(txt_dir, excel_path):
    wb = load_workbook(excel_path, data_only=True)
    updated_count = 0

    for txt_file in os.listdir(txt_dir):
        if not txt_file.endswith(".txt"):
            continue

        sheet_name = os.path.splitext(txt_file)[0]
        if sheet_name not in wb.sheetnames:
            continue

        txt_path = os.path.join(txt_dir, txt_file)
        with open(txt_path, "r", encoding="utf-8-sig") as f:
            content = f.read()

        ws = wb[sheet_name]
        ws.cell(row=6, column=2, value=content)
        updated_count += 1

    wb.save(excel_path)
    return updated_count

# Streamlitタブ内処理
with tab7:
    st.markdown(
        """
        ### STEP7: エクセルコメント書き換え
        <span style='font-size:10pt;'>分析済みテキスト（LR用）をExcelファイルに転記します。</span>
        """,
        unsafe_allow_html=True
    )

    # テキストフォルダのベースパス（session_stateに保持されていると仮定）
    if "base_dir" in st.session_state:
        base_dir = st.session_state["base_dir"]
        txt_dir = os.path.join(base_dir, "4 output for LR")

        # Excelファイルアップロード
        uploaded_file = st.file_uploader("Excelファイルをアップロードしてください", type=["xlsx"])

        if uploaded_file is not None:
            # 一時ファイルとして保存
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded_file.read())
                temp_excel_path = tmp.name

            if st.button("Excelファイルに書き込みを実行", key="commentback_excel_write"):
                try:
                    updated_count = write_txt_to_excel(txt_dir, temp_excel_path)

                    # 書き込み後のファイルをダウンロードリンクとして提供
                    with open(temp_excel_path, "rb") as f:
                        bytes_data = f.read()
                    st.success(f"✅ {updated_count} 件の内容をExcelファイルに転記しました！")
                    st.download_button(
                        label="🔽 編集済みExcelファイルをダウンロード",
                        data=bytes_data,
                        file_name="commented_output.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except Exception as e:
                    st.error(f"Excel書き込み中にエラーが発生しました：{e}")
        else:
            st.info("Excelファイルをアップロードしてください。")
    else:
        st.warning("STEP1が完了していないため、フォルダパスが取得できません。")
